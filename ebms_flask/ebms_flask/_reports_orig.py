"""Reports generation service with Excel export support.

Generates professional procurement reports and exports to Excel format.
"""
from datetime import datetime, timedelta
from io import BytesIO
import json
from flask import current_app
from sqlalchemy import func
from app.extensions import db
from app.models.procurement import Procurement
from app.models.submission import Submission
from app.models.bidder import Bidder
from app.models.evaluation import Evaluation
from app.models.complaint import Complaint
from app.models.audit import AuditLog
from app.utils.audit_enhanced import log_report_export


class ReportsService:
    """Service for generating professional reports."""
    
    @staticmethod
    def generate_bidder_participation_report(filters=None):
        """Generate bidder participation report.
        
        Args:
            filters: Dict with optional filters:
                - procurement_id: Specific procurement
                - bidder_id: Specific bidder
                - status: Submission status
                - start_date: Date range start
                - end_date: Date range end
                
        Returns:
            List of report rows as dicts
        """
        filters = filters or {}
        
        # Base query
        query = db.session.query(
            Procurement.id.label('procurement_id'),
            Procurement.tender_number,
            Procurement.title,
            Bidder.id.label('bidder_id'),
            Bidder.company_name,
            Submission.submitted_at,
            Submission.status,
        ).join(
            Submission, Procurement.id == Submission.procurement_id
        ).join(
            Bidder, Submission.bidder_id == Bidder.id
        )
        
        # Apply filters
        if filters.get('procurement_id'):
            query = query.filter(Procurement.id == filters['procurement_id'])
        
        if filters.get('bidder_id'):
            query = query.filter(Bidder.id == filters['bidder_id'])
        
        if filters.get('status'):
            query = query.filter(Submission.status == filters['status'])
        
        if filters.get('start_date'):
            query = query.filter(Submission.submitted_at >= filters['start_date'])
        
        if filters.get('end_date'):
            query = query.filter(Submission.submitted_at <= filters['end_date'])
        
        # Get results
        results = query.order_by(Procurement.tender_number, Bidder.company_name).all()
        
        # Format as dicts (keys match the bidder participation report template)
        rows = []
        for proc_id, tender_num, title, bidder_id, company, submit_date, status in results:
            # Representative evaluation score for this bidder on this procurement
            evaluation = Evaluation.query.filter(
                Evaluation.procurement_id == proc_id,
                Evaluation.bidder_id == bidder_id,
                Evaluation.score.isnot(None),
            ).order_by(Evaluation.id.desc()).first()

            evaluation_score = None
            if evaluation is not None:
                evaluation_score = float(evaluation.consensus_score or evaluation.score)

            # Has an award been published with this bidder as the winner?
            from app.models.award import Award
            award_status = 'awarded' if Award.query.filter_by(
                procurement_id=proc_id, winning_bidder_id=bidder_id
            ).first() else None

            rows.append({
                'procurement_title': title,
                'tender_number': tender_num,
                'bidder_name': company,
                'submission_date': submit_date,
                'status': status,
                'evaluation_score': evaluation_score,
                'award_status': award_status,
            })
        
        return rows

    @staticmethod
    def generate_procurement_summary_report(filters=None):
        """Generate procurement summary report.
        
        Args:
            filters: Dict with optional filters
            
        Returns:
            List of report rows
        """
        filters = filters or {}
        
        query = Procurement.query
        
        if filters.get('status'):
            query = query.filter(Procurement.status == filters['status'])
        
        if filters.get('category'):
            query = query.filter(Procurement.category == filters['category'])
        
        procurements = query.all()
        
        rows = []
        for proc in procurements:
            submission_count = Submission.query.filter_by(
                procurement_id=proc.id,
                status='submitted'
            ).count()
            
            evaluation_count = Evaluation.query.filter_by(
                procurement_id=proc.id
            ).distinct(Evaluation.bidder_id).count()
            
            rows.append({
                'Tender Number': proc.tender_number,
                'Title': proc.title,
                'Category': proc.category.replace('_', ' ').title(),
                'Method': proc.method.replace('_', ' ').title(),
                'Status': proc.status.replace('_', ' ').title(),
                'Estimated Value': f"{float(proc.estimated_value):,.2f}",
                'Created Date': proc.created_at.strftime('%Y-%m-%d') if proc.created_at else 'N/A',
                'Submission Deadline': proc.submission_deadline.strftime('%Y-%m-%d %H:%M') if proc.submission_deadline else 'TBA',
                'Submissions Received': submission_count,
                'Bidders Evaluated': evaluation_count,
            })
        
        return rows

    @staticmethod
    def generate_audit_report(filters=None, limit=500):
        """Generate audit trail report.
        
        Args:
            filters: Dict with optional filters (user_id, action, entity_type)
            limit: Maximum entries
            
        Returns:
            List of audit entries
        """
        filters = filters or {}
        
        query = AuditLog.query
        
        if filters.get('user_id'):
            query = query.filter(AuditLog.user_id == filters['user_id'])
        
        if filters.get('action'):
            query = query.filter(AuditLog.action.like(f"%{filters['action']}%"))
        
        if filters.get('entity_type'):
            query = query.filter(AuditLog.entity_type == filters['entity_type'])
        
        if filters.get('start_date'):
            query = query.filter(AuditLog.created_at >= filters['start_date'])
        
        if filters.get('end_date'):
            query = query.filter(AuditLog.created_at <= filters['end_date'])
        
        audit_logs = query.order_by(AuditLog.created_at.desc()).limit(limit).all()
        
        rows = []
        for log in audit_logs:
            rows.append({
                'Timestamp': log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'User': log.user.username if log.user else 'System',
                'Action': log.action,
                'Entity': log.entity_type,
                'Entity ID': log.entity_id or '-',
                'Reason': log.reason or '-',
                'IP Address': log.ip_address or '-',
            })
        
        return rows

    @staticmethod
    def generate_complaints_report(filters=None):
        """Generate complaints report.
        
        Args:
            filters: Optional filters
            
        Returns:
            List of complaint rows
        """
        filters = filters or {}
        
        query = Complaint.query
        
        if filters.get('procurement_id'):
            query = query.filter(Complaint.procurement_id == filters['procurement_id'])
        
        if filters.get('status'):
            query = query.filter(Complaint.status == filters['status'])
        
        if filters.get('start_date'):
            query = query.filter(Complaint.created_at >= filters['start_date'])
        
        if filters.get('end_date'):
            query = query.filter(Complaint.created_at <= filters['end_date'])
        
        complaints = query.order_by(Complaint.created_at.desc()).all()
        
        rows = []
        for complaint in complaints:
            rows.append({
                'Procurement': complaint.procurement.tender_number if complaint.procurement else 'N/A',
                'Bidder': complaint.bidder.company_name if complaint.bidder else 'N/A',
                'Date Lodged': complaint.created_at.strftime('%Y-%m-%d %H:%M') if complaint.created_at else 'N/A',
                'Status': complaint.status.replace('_', ' ').title(),
                'Grounds': complaint.grounds[:100] + '...' if complaint.grounds and len(complaint.grounds) > 100 else complaint.grounds or '',
                'Decision': complaint.decision[:50] + '...' if complaint.decision and len(complaint.decision) > 50 else complaint.decision or 'Pending',
                'Resolved By': complaint.resolved_by.full_name() if complaint.resolved_by else '-',
                'Resolved Date': complaint.resolved_at.strftime('%Y-%m-%d') if complaint.resolved_at else '-',
            })
        
        return rows


class ExcelExportService:
    """Service for exporting reports to Excel format."""
    
    @staticmethod
    def export_to_excel(report_data, report_title="Report", sheet_name="Data"):
        """Export report data to Excel.
        
        Args:
            report_data: List of dicts (rows)
            report_title: Title for the report
            sheet_name: Sheet name
            
        Returns:
            BytesIO object containing Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            current_app.logger.error("openpyxl not installed. Cannot export to Excel.")
            return None
        
        if not report_data:
            return None
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write title
        ws.append([report_title])
        title_cell = ws['A1']
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # Write timestamp
        ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws['A2'].font = Font(italic=True, size=10)
        
        # Write headers
        if report_data:
            headers = list(report_data[0].keys())
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Write data
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row_num, row_data in enumerate(report_data, 5):
                for col_num, value in enumerate(row_data.values(), 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = str(value) if value is not None else ""
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                max_length = max(
                    len(header),
                    max((len(str(row.get(header, ""))) for row in report_data), default=0)
                )
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

    @staticmethod
    def export_bidder_participation_report(filters=None):
        """Export bidder participation report to Excel.
        
        Args:
            filters: Report filters
            
        Returns:
            BytesIO object
        """
        data = ReportsService.generate_bidder_participation_report(filters)
        
        # Log export
        log_report_export(
            report_type='bidder_participation',
            filters=filters,
            format_type='excel'
        )
        
        return ExcelExportService.export_to_excel(
            data,
            report_title="Bidder Participation Report",
            sheet_name="Participations"
        )

    @staticmethod
    def export_procurement_summary_report(filters=None):
        """Export procurement summary report to Excel.
        
        Args:
            filters: Report filters
            
        Returns:
            BytesIO object
        """
        data = ReportsService.generate_procurement_summary_report(filters)
        
        log_report_export(
            report_type='procurement_summary',
            filters=filters,
            format_type='excel'
        )
        
        return ExcelExportService.export_to_excel(
            data,
            report_title="Procurement Summary Report",
            sheet_name="Procurements"
        )

    @staticmethod
    def export_audit_report(filters=None):
        """Export audit report to Excel.
        
        Args:
            filters: Report filters
            
        Returns:
            BytesIO object
        """
        data = ReportsService.generate_audit_report(filters)
        
        log_report_export(
            report_type='audit_trail',
            filters=filters,
            format_type='excel'
        )
        
        return ExcelExportService.export_to_excel(
            data,
            report_title="Audit Trail Report",
            sheet_name="Audit Log"
        )

    @staticmethod
    def export_complaints_report(filters=None):
        """Export complaints report to Excel.
        
        Args:
            filters: Report filters
            
        Returns:
            BytesIO object
        """
        data = ReportsService.generate_complaints_report(filters)
        
        log_report_export(
            report_type='complaints',
            filters=filters,
            format_type='excel'
        )
        
        return ExcelExportService.export_to_excel(
            data,
            report_title="Complaints Report",
            sheet_name="Complaints"
        )
