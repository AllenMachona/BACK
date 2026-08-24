"""Reports generation service with Excel export support.

Generates professional procurement reports and exports to Excel format.
"""
from datetime import datetime, timedelta
from io import BytesIO
import json
from flask import current_app
from sqlalchemy import func
from app.extensions import db
from app.models.procurement import Procurement
from app.models.submission import Submission
from app.models.bidder import Bidder
from app.models.evaluation import Evaluation
from app.models.complaint import Complaint
from app.models.audit import AuditLog
from app.models.award import Award
from app.models.communication import Communication
from app.models.payment import BidderPayment, BidderDocumentAccess
from app.models.budget_entry import BudgetEntry
from app.models.message import Message, MessageRecipient, MessageAttachment
from app.models.bidder_compliance import BidderComplianceDocument
from app.models.bidder_performance import BidderPerformance
from app.models.history import ProcurementHistory
from app.models.committee import CommitteeMember
from app.models.request import FormDRequest, FormERequest, FormDERequest
from app.models.user import User
from app.utils.audit_enhanced import log_report_export


def _status_label(value):
    """Convert a raw status code into a human-friendly label."""
    if not value:
        return '-'
    return value.replace('_', ' ').title()


class ReportsService:
    """Service for generating professional reports."""
    
    @staticmethod
    def generate_bidder_participation_report(filters=None):
        """Generate bidder participation report.
        
        Args:
            filters: Dict with optional filters:
                - procurement_id: Specific procurement
                - bidder_id: Specific bidder
                - status: Submission status
                - start_date: Date range start
                - end_date: Date range end
                
        Returns:
            List of report rows as dicts
        """
        filters = filters or {}
        
        # Base query
        query = db.session.query(
            Procurement.id.label('procurement_id'),
            Procurement.tender_number,
            Procurement.title,
            Bidder.id.label('bidder_id'),
            Bidder.company_name,
            Submission.submitted_at,
            Submission.status,
        ).join(
            Submission, Procurement.id == Submission.procurement_id
        ).join(
            Bidder, Submission.bidder_id == Bidder.id
        )
        
        # Apply filters
        if filters.get('procurement_id'):
            query = query.filter(Procurement.id == filters['procurement_id'])
        
        if filters.get('bidder_id'):
            query = query.filter(Bidder.id == filters['bidder_id'])
        
        if filters.get('status'):
            query = query.filter(Submission.status == filters['status'])
        
        if filters.get('start_date'):
            query = query.filter(Submission.submitted_at >= filters['start_date'])
        
        if filters.get('end_date'):
            query = query.filter(Submission.submitted_at <= filters['end_date'])
        
        # Get results
        results = query.order_by(Procurement.tender_number, Bidder.company_name).all()
        
        # Format as dicts (keys match the bidder participation report template)
        rows = []
        for proc_id, tender_num, title, bidder_id, company, submit_date, status in results:
            # Representative evaluation score for this bidder on this procurement
            evaluation = Evaluation.query.filter(
                Evaluation.procurement_id == proc_id,
                Evaluation.bidder_id == bidder_id,
                Evaluation.score.isnot(None),
            ).order_by(Evaluation.id.desc()).first()

            evaluation_score = None
            if evaluation is not None:
                evaluation_score = float(evaluation.consensus_score or evaluation.score)

            # Has an award been published with this bidder as the winner?
            from app.models.award import Award
            award_status = 'awarded' if Award.query.filter_by(
                procurement_id=proc_id, winning_bidder_id=bidder_id
            ).first() else None

            rows.append({
                'procurement_title': title,
                'tender_number': tender_num,
                'bidder_name': company,
                'submission_date': submit_date,
                'status': status,
                'evaluation_score': evaluation_score,
                'award_status': award_status,
            })
        
        return rows

    @staticmethod
    def generate_procurement_summary_report(filters=None):
        """Generate procurement summary report.
        
        Args:
            filters: Dict with optional filters
            
        Returns:
            List of report rows
        """
        filters = filters or {}
        
        query = Procurement.query
        
        if filters.get('status'):
            query = query.filter(Procurement.status == filters['status'])
        
        if filters.get('category'):
            query = query.filter(Procurement.category == filters['category'])
        
        procurements = query.all()
        
        rows = []
        for proc in procurements:
            submission_count = Submission.query.filter_by(
                procurement_id=proc.id,
                status='submitted'
            ).count()
            
            evaluation_count = Evaluation.query.filter_by(
                procurement_id=proc.id
            ).distinct(Evaluation.bidder_id).count()
            
            rows.append({
                'Tender Number': proc.tender_number,
                'Title': proc.title,
                'Category': proc.category.replace('_', ' ').title(),
                'Method': proc.method.replace('_', ' ').title(),
                'Status': proc.status.replace('_', ' ').title(),
                'Estimated Value': f"{float(proc.estimated_value):,.2f}",
                'Created Date': proc.created_at.strftime('%Y-%m-%d') if proc.created_at else 'N/A',
                'Submission Deadline': proc.submission_deadline.strftime('%Y-%m-%d %H:%M') if proc.submission_deadline else 'TBA',
                'Submissions Received': submission_count,
                'Bidders Evaluated': evaluation_count,
            })
        
        return rows

    @staticmethod
    def generate_audit_report(filters=None, limit=500):
        """Generate audit trail report.
        
        Args:
            filters: Dict with optional filters (user_id, action, entity_type)
            limit: Maximum entries
            
        Returns:
            List of audit entries
        """
        filters = filters or {}
        
        query = AuditLog.query
        
        if filters.get('user_id'):
            query = query.filter(AuditLog.user_id == filters['user_id'])
        
        if filters.get('action'):
            query = query.filter(AuditLog.action.like(f"%{filters['action']}%"))
        
        if filters.get('entity_type'):
            query = query.filter(AuditLog.entity_type == filters['entity_type'])
        
        if filters.get('start_date'):
            query = query.filter(AuditLog.created_at >= filters['start_date'])
        
        if filters.get('end_date'):
            query = query.filter(AuditLog.created_at <= filters['end_date'])
        
        audit_logs = query.order_by(AuditLog.created_at.desc()).limit(limit).all()
        
        rows = []
        for log in audit_logs:
            rows.append({
                'Timestamp': log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'User': log.user.username if log.user else 'System',
                'Action': log.action,
                'Entity': log.entity_type,
                'Entity ID': log.entity_id or '-',
                'Reason': log.reason or '-',
                'IP Address': log.ip_address or '-',
            })
        
        return rows

    @staticmethod
    def generate_complaints_report(filters=None):
        """Generate complaints report.
        
        Args:
            filters: Optional filters
            
        Returns:
            List of complaint rows
        """
        filters = filters or {}
        
        query = Complaint.query
        
        if filters.get('procurement_id'):
            query = query.filter(Complaint.procurement_id == filters['procurement_id'])
        
        if filters.get('status'):
            query = query.filter(Complaint.status == filters['status'])
        
        if filters.get('start_date'):
            query = query.filter(Complaint.created_at >= filters['start_date'])
        
        if filters.get('end_date'):
            query = query.filter(Complaint.created_at <= filters['end_date'])
        
        complaints = query.order_by(Complaint.created_at.desc()).all()
        
        rows = []
        for complaint in complaints:
            rows.append({
                'Procurement': complaint.procurement.tender_number if complaint.procurement else 'N/A',
                'Bidder': complaint.bidder.company_name if complaint.bidder else 'N/A',
                'Date Lodged': complaint.created_at.strftime('%Y-%m-%d %H:%M') if complaint.created_at else 'N/A',
                'Status': complaint.status.replace('_', ' ').title(),
                'Grounds': complaint.grounds[:100] + '...' if complaint.grounds and len(complaint.grounds) > 100 else complaint.grounds or '',
                'Decision': complaint.decision[:50] + '...' if complaint.decision and len(complaint.decision) > 50 else complaint.decision or 'Pending',
                'Resolved By': complaint.resolved_by.full_name() if complaint.resolved_by else '-',
                'Resolved Date': complaint.resolved_at.strftime('%Y-%m-%d') if complaint.resolved_at else '-',
            })
        
        return rows

    # ------------------------------------------------------------------
    # Tender Register
    # ------------------------------------------------------------------
    @staticmethod
    def generate_tender_register_report(filters=None):
        """Full register of all procurements/tenders with lifecycle metrics."""
        filters = filters or {}
        query = Procurement.query

        if filters.get('status'):
            query = query.filter(Procurement.status == filters['status'])
        if filters.get('category'):
            query = query.filter(Procurement.category == filters['category'])
        if filters.get('method'):
            query = query.filter(Procurement.method == filters['method'])
        if filters.get('start_date'):
            query = query.filter(Procurement.created_at >= filters['start_date'])
        if filters.get('end_date'):
            query = query.filter(Procurement.created_at <= filters['end_date'])

        rows = []
        for proc in query.order_by(Procurement.created_at.desc()).all():
            sub_count = proc.submissions.filter_by(status='submitted').count()
            eval_count = db.session.query(Evaluation.bidder_id).filter(
                Evaluation.procurement_id == proc.id
            ).distinct().count()
            rows.append({
                'Tender Number': proc.tender_number,
                'Title': proc.title,
                'Category': proc.category,
                'Method': proc.method,
                'Procurement Entity': proc.procurement_entity or '-',
                'Status': proc.status,
                'Estimated Value': f'{float(proc.estimated_value or 0):,.2f}',
                'Submission Deadline': proc.submission_deadline.strftime('%Y-%m-%d %H:%M') if proc.submission_deadline else 'TBA',
                'Tender Fee': f'{float(proc.tender_fee or 0):,.2f}',
                'Created Date': proc.created_at.strftime('%Y-%m-%d') if proc.created_at else 'N/A',
                'Submissions': sub_count,
                'Bidders Evaluated': eval_count,
            })
        return rows

    @staticmethod
    def export_tender_register_report(filters=None):
        data = ReportsService.generate_tender_register_report(filters)
        log_report_export(report_type='tender_register', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Tender Register Report", "Tender Register")

    # ------------------------------------------------------------------
    # Awards & Contracts
    # ------------------------------------------------------------------
    @staticmethod
    def generate_award_report(filters=None):
        """Awards and contract conclusion status."""
        filters = filters or {}
        query = Award.query
        if filters.get('procurement_id'):
            query = query.filter(Award.procurement_id == filters['procurement_id'])
        if filters.get('contract_status') == 'concluded':
            query = query.filter(Award.contract_concluded.is_(True))
        elif filters.get('contract_status') == 'cooling_off':
            query = query.filter(Award.contract_concluded.is_(False))
        if filters.get('start_date'):
            query = query.filter(Award.decision_date >= filters['start_date'])
        if filters.get('end_date'):
            query = query.filter(Award.decision_date <= filters['end_date'])

        rows = []
        for award in query.order_by(Award.decision_date.desc()).all():
            proc = award.procurement
            bidder = award.winning_bidder
            rows.append({
                'Tender Number': proc.tender_number if proc else '-',
                'Title': proc.title if proc else '-',
                'Winning Bidder': bidder.company_name if bidder else '-',
                'Award Value': f'{float(proc.estimated_value or 0):,.2f}' if proc else '-',
                'Decision Date': award.decision_date.strftime('%Y-%m-%d') if award.decision_date else '-',
                'Cooling-off Expiry': award.cooling_off_expiry.strftime('%Y-%m-%d') if award.cooling_off_expiry else '-',
                'Cooling-off Active': 'Yes' if award.cooling_off_active() else 'No',
                'Contract Concluded': 'Yes' if award.contract_concluded else 'No',
                'Concluded At': award.contract_concluded_at.strftime('%Y-%m-%d') if award.contract_concluded_at else '-',
            })
        return rows

    @staticmethod
    def export_award_report(filters=None):
        data = ReportsService.generate_award_report(filters)
        log_report_export(report_type='awards', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Awards & Contracts Report", "Awards")

    # ------------------------------------------------------------------
    # Evaluations
    # ------------------------------------------------------------------
    @staticmethod
    def generate_evaluation_report(filters=None):
        """Evaluation outcomes across procurements."""
        filters = filters or {}
        query = Evaluation.query
        if filters.get('procurement_id'):
            query = query.filter(Evaluation.procurement_id == filters['procurement_id'])
        if filters.get('stage'):
            query = query.filter(Evaluation.evaluation_stage == filters['stage'])
        if filters.get('passed') in (True, 'true', 'True'):
            query = query.filter(Evaluation.passed.is_(True))
        elif filters.get('passed') in (False, 'false', 'False'):
            query = query.filter(Evaluation.passed.is_(False))
        if filters.get('start_date'):
            query = query.filter(Evaluation.created_at >= filters['start_date'])
        if filters.get('end_date'):
            query = query.filter(Evaluation.created_at <= filters['end_date'])

        rows = []
        for ev in query.order_by(Evaluation.created_at.desc()).all():
            proc = ev.procurement
            bidder = ev.bidder
            evaluator = ev.evaluator
            score = float(ev.score) if ev.score is not None else None
            cons = float(ev.consensus_score) if ev.consensus_score is not None else None
            rows.append({
                'Tender Number': proc.tender_number if proc else '-',
                'Procurement': proc.title if proc else '-',
                'Bidder': bidder.company_name if bidder else '-',
                'Evaluator': evaluator.full_name() if evaluator else '-',
                'Stage': _status_label(ev.evaluation_stage),
                'Score': score if score is not None else '-',
                'Consensus Score': cons if cons is not None else '-',
                'Passed': 'Yes' if ev.passed else ('No' if ev.passed is False else '-'),
                'Eliminated': 'Yes' if ev.eliminated else 'No',
                'Date Reviewed': ev.created_at.strftime('%Y-%m-%d') if ev.created_at else '-',
            })
        return rows

    @staticmethod
    def export_evaluation_report(filters=None):
        data = ReportsService.generate_evaluation_report(filters)
        log_report_export(report_type='evaluations', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Evaluation Results Report", "Evaluations")

    # ------------------------------------------------------------------
    # Payment Verification
    # ------------------------------------------------------------------
    @staticmethod
    def generate_payment_report(filters=None):
        """Bidder tender-document payments and their verification status."""
        filters = filters or {}
        query = BidderPayment.query
        if filters.get('procurement_id'):
            query = query.filter(BidderPayment.procurement_id == filters['procurement_id'])
        if filters.get('status'):
            query = query.filter(BidderPayment.status == filters['status'])
        if filters.get('start_date'):
            query = query.filter(BidderPayment.submitted_at >= filters['start_date'])
        if filters.get('end_date'):
            query = query.filter(BidderPayment.submitted_at <= filters['end_date'])

        rows = []
        for pay in query.order_by(BidderPayment.submitted_at.desc()).all():
            proc = pay.procurement
            bidder = pay.bidder
            reviewer = pay.reviewed_by
            note = pay.notes or ''
            rows.append({
                'Tender Number': proc.tender_number if proc else '-',
                'Bidder': bidder.company_name if bidder else '-',
                'Payment Reference': pay.payment_reference,
                'Amount': f'{float(pay.amount or 0):,.2f}',
                'Status': pay.status,
                'Submitted At': pay.submitted_at.strftime('%Y-%m-%d %H:%M') if pay.submitted_at else '-',
                'Reviewed By': reviewer.full_name() if reviewer else '-',
                'Reviewed At': pay.reviewed_at.strftime('%Y-%m-%d %H:%M') if pay.reviewed_at else '-',
                'Notes': (note[:60] + '...') if len(note) > 60 else note,
            })
        return rows

    @staticmethod
    def export_payment_report(filters=None):
        data = ReportsService.generate_payment_report(filters)
        log_report_export(report_type='payments', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Payment Verification Report", "Payments")

    # ------------------------------------------------------------------
    # Bidder Registry
    # ------------------------------------------------------------------
    @staticmethod
    def generate_bidder_registry_report(filters=None):
        """Registered bidder companies and their registration status."""
        filters = filters or {}
        bidders = Bidder.query.order_by(Bidder.company_name).all()
        rows = []
        for bidder in bidders:
            reg = bidder.registration_status()
            if filters.get('status') and reg != filters['status']:
                continue
            if filters.get('verified') == 'yes' and not bidder.verified:
                continue
            if filters.get('verified') == 'no' and bidder.verified:
                continue
            rows.append({
                'Company Name': bidder.company_name,
                'Registration Number': bidder.ppra_registration_number or '-',
                'Grade': bidder.ppra_grade or '-',
                'Category': bidder.category or '-',
                'Registration Status': reg,
                'Verified': 'Yes' if bidder.verified else 'No',
                'Registered At': bidder.registered_at.strftime('%Y-%m-%d') if bidder.registered_at else '-',
                'Registration Expiry': bidder.registration_expiry.strftime('%Y-%m-%d') if bidder.registration_expiry else '-',
                'Email': bidder.contact_email or '-',
                'Phone': bidder.contact_phone or '-',
            })
        return rows

    @staticmethod
    def export_bidder_registry_report(filters=None):
        data = ReportsService.generate_bidder_registry_report(filters)
        log_report_export(report_type='bidder_registry', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Bidder Registry Report", "Bidders")

    # ------------------------------------------------------------------
    # User & Account
    # ------------------------------------------------------------------
    @staticmethod
    def generate_user_account_report(filters=None):
        """All system accounts / users with roles and access status."""
        filters = filters or {}
        users = User.query.order_by(User.last_name, User.first_name).all()
        rows = []
        for user in users:
            role = user.role
            if filters.get('role') and (not role or role.code != filters['role']):
                continue
            if filters.get('status') == 'active' and not user.is_active:
                continue
            if filters.get('status') == 'inactive' and user.is_active:
                continue
            if filters.get('start_date') and (not user.created_at or user.created_at < filters['start_date']):
                continue
            if filters.get('end_date') and user.created_at and user.created_at > filters['end_date']:
                continue
            rows.append({
                'Username': user.username,
                'Full Name': f'{user.first_name} {user.last_name}'.strip(),
                'Role': role.name if role else '-',
                'Role Code': role.code if role else '-',
                'Department': user.department or '-',
                'Email': user.email,
                'Status': 'active' if user.is_active else 'inactive',
                'MFA Enabled': 'Yes' if user.mfa_enabled else 'No',
                'Last Login': user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never',
                'Created At': user.created_at.strftime('%Y-%m-%d') if user.created_at else '-',
            })
        return rows

    @staticmethod
    def export_user_account_report(filters=None):
        data = ReportsService.generate_user_account_report(filters)
        log_report_export(report_type='users', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "User & Account Report", "Users")

    # ------------------------------------------------------------------
    # Communication & Clarifications
    # ------------------------------------------------------------------
    @staticmethod
    def generate_communication_report(filters=None):
        """All clarifications / communications / notices issued to bidders."""
        filters = filters or {}
        query = Communication.query
        if filters.get('procurement_id'):
            query = query.filter(Communication.procurement_id == filters['procurement_id'])
        if filters.get('type'):
            query = query.filter(Communication.type == filters['type'])
        if filters.get('visibility'):
            query = query.filter(Communication.visibility_type == filters['visibility'])
        if filters.get('start_date'):
            query = query.filter(Communication.created_at >= filters['start_date'])
        if filters.get('end_date'):
            query = query.filter(Communication.created_at <= filters['end_date'])

        rows = []
        for comm in query.order_by(Communication.created_at.desc()).all():
            proc = comm.procurement
            if comm.from_user:
                source = comm.from_user.full_name()
            elif comm.from_bidder:
                source = comm.from_bidder.company_name
            else:
                source = 'System'
            content = comm.content or ''
            rows.append({
                'Tender Number': proc.tender_number if proc else '-',
                'Type': _status_label(getattr(comm, 'type', 'communication')),
                'From': source,
                'Visibility': comm.visibility_type or 'public',
                'Public': 'Yes' if comm.is_public else 'No',
                'Attachment': 'Yes' if comm.original_filename else 'No',
                'Content Summary': (content[:60] + '...') if len(content) > 60 else content,
                'Posted At': comm.created_at.strftime('%Y-%m-%d %H:%M') if comm.created_at else '-',
            })
        return rows

    @staticmethod
    def export_communication_report(filters=None):
        data = ReportsService.generate_communication_report(filters)
        log_report_export(report_type='communications', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Communication & Clarification Report", "Communications")

    # ------------------------------------------------------------------
    # Request Pipeline
    # ------------------------------------------------------------------
    @staticmethod
    def generate_request_pipeline_report(filters=None):
        """Procurement request lifecycle (Form D / Form E / combined requests)."""
        filters = filters or {}
        form_map = [
            (FormDRequest, 'Form D'),
            (FormERequest, 'Form E'),
            (FormDERequest, 'Form D & E'),
        ]
        rows = []
        for model_cls, form_label in form_map:
            query = model_cls.query
            if filters.get('status'):
                query = query.filter(model_cls.status == filters['status'])
            if filters.get('form_type') and filters['form_type'] != form_label:
                continue
            if filters.get('start_date'):
                query = query.filter(model_cls.created_at >= filters['start_date'])
            if filters.get('end_date'):
                query = query.filter(model_cls.created_at <= filters['end_date'])
            for req in query.order_by(model_cls.created_at.desc()).all():
                requester = req.requester
                budget = req.estimated_value if hasattr(req, 'estimated_value') else getattr(req, 'budget_allocated', None)
                decision = req.converted_at or req.rejected_at
                department = requester.department if requester and requester.department else getattr(req, 'department', '-') or '-'
                rows.append({
                    'Request No': f'{form_label} #{req.id}',
                    'Form Type': form_label,
                    'Title': req.title,
                    'Requester': requester.full_name() if requester else '-',
                    'Department': department,
                    'Budget': f'{float(budget or 0):,.2f}' if budget is not None else '-',
                    'Status': req.status,
                    'Linked Procurement': req.procurement.tender_number if req.procurement else '-',
                    'Submitted At': req.created_at.strftime('%Y-%m-%d %H:%M') if req.created_at else '-',
                    'Converted / Rejected On': decision.strftime('%Y-%m-%d') if decision else '-',
                })
        rows.sort(key=lambda r: r.get('Submitted At') or '', reverse=True)
        return rows

    @staticmethod
    def export_request_pipeline_report(filters=None):
        data = ReportsService.generate_request_pipeline_report(filters)
        log_report_export(report_type='requests', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Request Pipeline Report", "Requests")

    # ------------------------------------------------------------------
    # Budget & Expenditure
    # ------------------------------------------------------------------
    @staticmethod
    def generate_budget_report(filters=None):
        """Budget ledger / spend against procurements."""
        filters = filters or {}
        query = BudgetEntry.query
        if filters.get('procurement_id'):
            query = query.filter(BudgetEntry.procurement_id == filters['procurement_id'])
        if filters.get('entry_type'):
            query = query.filter(BudgetEntry.entry_type == filters['entry_type'])
        if filters.get('start_date'):
            query = query.filter(BudgetEntry.entry_date >= filters['start_date'])
        if filters.get('end_date'):
            query = query.filter(BudgetEntry.entry_date <= filters['end_date'])

        rows = []
        for entry in query.order_by(BudgetEntry.entry_date.desc()).all():
            proc = entry.procurement
            created_by = entry.created_by
            rows.append({
                'Tender Number': proc.tender_number if proc else '-',
                'Procurement': proc.title if proc else '-',
                'Entry Type': entry.entry_type,
                'Description': entry.description,
                'Amount': f'{float(entry.amount or 0):,.2f}',
                'Reference': entry.reference or '-',
                'Entry Date': entry.entry_date.strftime('%Y-%m-%d') if entry.entry_date else '-',
                'Created By': created_by.full_name() if created_by else '-',
            })
        return rows

    @staticmethod
    def export_budget_report(filters=None):
        data = ReportsService.generate_budget_report(filters)
        log_report_export(report_type='budget', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Budget & Expenditure Report", "Budget")

    # ------------------------------------------------------------------
    # Submission Activity
    # ------------------------------------------------------------------
    @staticmethod
    def generate_submission_activity_report(filters=None):
        """Submitted / withdrawn / replaced sealed bids."""
        filters = filters or {}
        query = db.session.query(
            Procurement.id.label('procurement_id'),
            Procurement.tender_number,
            Procurement.title,
            Bidder.company_name,
            Submission.envelope_type,
            Submission.version,
            Submission.status,
            Submission.submitted_at,
            Submission.receipt_code,
            Submission.original_filename,
            Submission.file_size_bytes,
        ).join(Submission, Procurement.id == Submission.procurement_id).join(Bidder, Submission.bidder_id == Bidder.id)

        if filters.get('procurement_id'):
            query = query.filter(Procurement.id == filters['procurement_id'])
        if filters.get('bidder_id'):
            query = query.filter(Bidder.id == filters['bidder_id'])
        if filters.get('status'):
            query = query.filter(Submission.status == filters['status'])
        if filters.get('start_date'):
            query = query.filter(Submission.submitted_at >= filters['start_date'])
        if filters.get('end_date'):
            query = query.filter(Submission.submitted_at <= filters['end_date'])

        rows = []
        for record in query.order_by(Submission.submitted_at.desc()).all():
            size = record.file_size_bytes or 0
            rows.append({
                'Tender Number': record.tender_number,
                'Bidder': record.company_name,
                'Envelope': record.envelope_type,
                'Version': record.version,
                'Status': record.status,
                'Submitted At': record.submitted_at.strftime('%Y-%m-%d %H:%M') if record.submitted_at else '-',
                'Receipt Code': record.receipt_code or '-',
                'Original File': record.original_filename or '-',
                'File Size (KB)': round(size / 1024.0, 2) if size else 0,
            })
        return rows

    @staticmethod
    def export_submission_activity_report(filters=None):
        data = ReportsService.generate_submission_activity_report(filters)
        log_report_export(report_type='submissions', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Submission Activity Report", "Submissions")

    # ------------------------------------------------------------------
    # Bidder Performance
    # ------------------------------------------------------------------
    @staticmethod
    def generate_bidder_performance_report(filters=None):
        """Performance reviews of awarded bidders per procurement."""
        filters = filters or {}
        query = BidderPerformance.query
        if filters.get('procurement_id'):
            query = query.filter(BidderPerformance.procurement_id == filters['procurement_id'])
        if filters.get('status'):
            query = query.filter(BidderPerformance.status == filters['status'])
        rows = []
        for review in query.order_by(BidderPerformance.reviewed_at.desc()).all():
            proc = review.procurement
            bidder = review.bidder
            reviewed_by = review.reviewed_by
            note = review.notes or ''
            rows.append({
                'Tender Number': proc.tender_number if proc else '-',
                'Bidder': bidder.company_name if bidder else '-',
                'Delivery Score': review.delivery_score,
                'Quality Score': review.quality_score,
                'Compliance Score': review.compliance_score,
                'Overall Score': float(review.overall_score) if review.overall_score is not None else '-',
                'Status': review.status,
                'Reviewed By': reviewed_by.full_name() if reviewed_by else '-',
                'Reviewed At': review.reviewed_at.strftime('%Y-%m-%d') if review.reviewed_at else '-',
                'Notes': (note[:60] + '...') if len(note) > 60 else note,
            })
        return rows

    @staticmethod
    def export_bidder_performance_report(filters=None):
        data = ReportsService.generate_bidder_performance_report(filters)
        log_report_export(report_type='bidder_performance', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Bidder Performance Report", "Performance")

    # ------------------------------------------------------------------
    # Messaging Activity
    # ------------------------------------------------------------------
    @staticmethod
    def generate_message_report(filters=None):
        """Internal messaging volume and delivery activity."""
        filters = filters or {}
        query = Message.query
        if filters.get('message_type'):
            query = query.filter(Message.message_type == filters['message_type'])
        if filters.get('sender_id'):
            query = query.filter(Message.sender_id == filters['sender_id'])
        if filters.get('start_date'):
            query = query.filter(Message.created_at >= filters['start_date'])
        if filters.get('end_date'):
            query = query.filter(Message.created_at <= filters['end_date'])

        rows = []
        for message in query.order_by(Message.created_at.desc()).all():
            sender = message.sender
            recipient_query = message.recipients.filter(
                MessageRecipient.user_id != message.sender_id
            )
            recipient_count = recipient_query.count()
            unread_count = recipient_query.filter(
                MessageRecipient.read_at.is_(None)
            ).count()
            rows.append({
                'Subject': message.subject,
                'Sender': sender.full_name() if sender else '-',
                'Message Type': message.message_type,
                'Recipients': recipient_count,
                'Unread': unread_count,
                'Attachments': len(message.attachments) if message.attachments else 0,
                'Sent At': message.created_at.strftime('%Y-%m-%d %H:%M') if message.created_at else '-',
            })
        return rows

    @staticmethod
    def export_message_report(filters=None):
        data = ReportsService.generate_message_report(filters)
        log_report_export(report_type='messages', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Messaging Activity Report", "Messages")

    # ------------------------------------------------------------------
    # Compliance Documents
    # ------------------------------------------------------------------
    @staticmethod
    def generate_compliance_report(filters=None):
        """Bidder registration compliance documents."""
        filters = filters or {}
        query = BidderComplianceDocument.query
        if filters.get('status'):
            query = query.filter(BidderComplianceDocument.status == filters['status'])
        rows = []
        for doc in query.order_by(BidderComplianceDocument.submitted_at.desc()).all():
            bidder = doc.bidder
            reviewer = doc.reviewed_by
            note = doc.review_notes or ''
            rows.append({
                'Company Name': bidder.company_name if bidder else '-',
                'File Name': doc.original_filename,
                'Status': doc.status,
                'Submitted At': doc.submitted_at.strftime('%Y-%m-%d %H:%M') if doc.submitted_at else '-',
                'Reviewed By': reviewer.full_name() if reviewer else '-',
                'Reviewed At': doc.reviewed_at.strftime('%Y-%m-%d') if doc.reviewed_at else '-',
                'Review Notes': (note[:60] + '...') if len(note) > 60 else note,
            })
        return rows

    @staticmethod
    def export_compliance_report(filters=None):
        data = ReportsService.generate_compliance_report(filters)
        log_report_export(report_type='compliance', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Compliance Documents Report", "Compliance")

    # ------------------------------------------------------------------
    # Procurement History (state transitions)
    # ------------------------------------------------------------------
    @staticmethod
    def generate_procurement_history_report(filters=None):
        """Lifecycle state changes for procurements.

        The history table is the primary source. Older deployments recorded
        transitions only in AuditLog, so unmatched status-change audit rows are
        included as a compatibility fallback.
        """
        filters = filters or {}
        query = ProcurementHistory.query
        if filters.get('procurement_id'):
            query = query.filter(ProcurementHistory.procurement_id == filters['procurement_id'])
        if filters.get('action'):
            query = query.filter(ProcurementHistory.action == filters['action'])
        if filters.get('start_date'):
            query = query.filter(ProcurementHistory.performed_at >= filters['start_date'])
        if filters.get('end_date'):
            query = query.filter(ProcurementHistory.performed_at <= filters['end_date'])

        history_entries = query.order_by(ProcurementHistory.performed_at.desc()).all()
        rows = []
        history_keys = {
            (entry.procurement_id, entry.previous_status, entry.new_status, entry.performed_by_id)
            for entry in history_entries
        }

        for entry in history_entries:
            proc = entry.procurement
            performed_by = entry.performed_by
            approved_by = entry.approved_by
            reason = entry.reason or ''
            rows.append({
                'Tender Number': proc.tender_number if proc else '-',
                'Action': entry.action,
                'Previous Status': entry.previous_status or '-',
                'New Status': entry.new_status or '-',
                'Performed By': performed_by.full_name() if performed_by else '-',
                'Performed At': entry.performed_at.strftime('%Y-%m-%d %H:%M') if entry.performed_at else '-',
                'Requires Approval': 'Yes' if entry.requires_approval else 'No',
                'Approved By': approved_by.full_name() if approved_by else '-',
                'Reason': (reason[:60] + '...') if len(reason) > 60 else reason,
            })

        audit_query = AuditLog.query.filter(
            AuditLog.entity_type == 'Procurement',
            AuditLog.action == 'PROCUREMENT_STATUS_CHANGED',
        )
        if filters.get('procurement_id'):
            audit_query = audit_query.filter(AuditLog.entity_id == filters['procurement_id'])
        if filters.get('start_date'):
            audit_query = audit_query.filter(AuditLog.created_at >= filters['start_date'])
        if filters.get('end_date'):
            audit_query = audit_query.filter(AuditLog.created_at <= filters['end_date'])

        for audit_entry in audit_query.order_by(AuditLog.created_at.desc()).all():
            try:
                previous_value = json.loads(audit_entry.previous_value or '{}')
                new_value = json.loads(audit_entry.new_value or '{}')
            except (TypeError, ValueError):
                previous_value, new_value = {}, {}
            previous_status = previous_value.get('status')
            new_status = new_value.get('status')
            audit_key = (audit_entry.entity_id, previous_status, new_status, audit_entry.user_id)
            if audit_key in history_keys:
                continue
            if filters.get('action') and filters['action'] not in ('status_changed', new_status):
                continue
            proc = Procurement.query.get(audit_entry.entity_id)
            user = User.query.get(audit_entry.user_id) if audit_entry.user_id else None
            reason = audit_entry.reason or ''
            rows.append({
                'Tender Number': proc.tender_number if proc else '-',
                'Action': 'status_changed',
                'Previous Status': previous_status or '-',
                'New Status': new_status or '-',
                'Performed By': user.full_name() if user else '-',
                'Performed At': audit_entry.created_at.strftime('%Y-%m-%d %H:%M') if audit_entry.created_at else '-',
                'Requires Approval': 'No',
                'Approved By': '-',
                'Reason': (reason[:60] + '...') if len(reason) > 60 else reason,
            })

        rows.sort(key=lambda row: row['Performed At'], reverse=True)
        return rows

    @staticmethod
    def export_procurement_history_report(filters=None):
        data = ReportsService.generate_procurement_history_report(filters)
        log_report_export(report_type='procurement_history', filters=filters, format_type='excel')
        return ExcelExportService.export_to_excel(data, "Procurement History Report", "History")


class ExcelExportService:
    """Service for exporting reports to Excel format."""
    
    @staticmethod
    def export_to_excel(report_data, report_title="Report", sheet_name="Data"):
        """Export report data to Excel.
        
        Args:
            report_data: List of dicts (rows)
            report_title: Title for the report
            sheet_name: Sheet name
            
        Returns:
            BytesIO object containing Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            current_app.logger.error("openpyxl not installed. Cannot export to Excel.")
            return None
        
        if not report_data:
            return None
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write title
        ws.append([report_title])
        title_cell = ws['A1']
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # Write timestamp
        ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws['A2'].font = Font(italic=True, size=10)
        
        # Write headers
        if report_data:
            headers = list(report_data[0].keys())
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Write data
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row_num, row_data in enumerate(report_data, 5):
                for col_num, value in enumerate(row_data.values(), 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = str(value) if value is not None else ""
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                max_length = max(
                    len(header),
                    max((len(str(row.get(header, ""))) for row in report_data), default=0)
                )
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

    @staticmethod
    def export_bidder_participation_report(filters=None):
        """Export bidder participation report to Excel.
        
        Args:
            filters: Report filters
            
        Returns:
            BytesIO object
        """
        data = ReportsService.generate_bidder_participation_report(filters)
        
        # Log export
        log_report_export(
            report_type='bidder_participation',
            filters=filters,
            format_type='excel'
        )
        
        return ExcelExportService.export_to_excel(
            data,
            report_title="Bidder Participation Report",
            sheet_name="Participations"
        )

    @staticmethod
    def export_procurement_summary_report(filters=None):
        """Export procurement summary report to Excel.
        
        Args:
            filters: Report filters
            
        Returns:
            BytesIO object
        """
        data = ReportsService.generate_procurement_summary_report(filters)
        
        log_report_export(
            report_type='procurement_summary',
            filters=filters,
            format_type='excel'
        )
        
        return ExcelExportService.export_to_excel(
            data,
            report_title="Procurement Summary Report",
            sheet_name="Procurements"
        )

    @staticmethod
    def export_audit_report(filters=None):
        """Export audit report to Excel.
        
        Args:
            filters: Report filters
            
        Returns:
            BytesIO object
        """
        data = ReportsService.generate_audit_report(filters)
        
        log_report_export(
            report_type='audit_trail',
            filters=filters,
            format_type='excel'
        )
        
        return ExcelExportService.export_to_excel(
            data,
            report_title="Audit Trail Report",
            sheet_name="Audit Log"
        )

    @staticmethod
    def export_complaints_report(filters=None):
        """Export complaints report to Excel.
        
        Args:
            filters: Report filters
            
        Returns:
            BytesIO object
        """
        data = ReportsService.generate_complaints_report(filters)
        
        log_report_export(
            report_type='complaints',
            filters=filters,
            format_type='excel'
        )
        
        return ExcelExportService.export_to_excel(
            data,
            report_title="Complaints Report",
            sheet_name="Complaints"
        )
